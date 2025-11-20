import json
from django.shortcuts import render, HttpResponse
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Sum, Q, F, Case, When, DecimalField
from django.utils import timezone
from datetime import datetime
from .forms import ReportFilterForm
from .models import SavedReport
from customers.models import Customer
from opportunities.models import Opportunity
from followups.models import Followup
from sales_team.models import TeamMember
from django.contrib.auth.models import User
import csv
import logging

logger = logging.getLogger(__name__)


def manager_required(view_func):
    """Decorador para verificar permisos de gerente/admin"""
    def wrapper(request, *args, **kwargs):
        if request.user.profile.role not in ('administrador', 'gerente'):
            raise PermissionDenied("No tienes permiso para acceder a reportes")
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
@manager_required
def report_view(request):
    """Vista principal de reportes"""
    form = ReportFilterForm(request.GET or None)
    report_data = None
    report_type = None
    report_title = None
    export_format = None
    
    if form.is_valid():
        report_type = form.cleaned_data.get('report_type')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        export_format = form.cleaned_data.get('export_format')
        
        # Generar reporte según tipo
        if report_type == 'sales':
            report_data, report_title = get_sales_report(start_date, end_date)
        elif report_type == 'customers':
            report_data, report_title = get_customers_report(
                start_date, end_date,
                form.cleaned_data.get('assigned_to')
            )
        elif report_type == 'sellers':
            report_data, report_title = get_sellers_report(start_date, end_date)
        elif report_type == 'followups':
            report_data, report_title = get_followups_report(
                start_date, end_date,
                form.cleaned_data.get('assigned_to')
            )
        elif report_type == 'opportunities':
            report_data, report_title = get_opportunities_report(
                start_date, end_date,
                form.cleaned_data.get('customer'),
                form.cleaned_data.get('assigned_to'),
                form.cleaned_data.get('opportunity_status')
            )
        elif report_type == 'pipeline':
            report_data, report_title = get_pipeline_report(
                start_date, end_date,
                form.cleaned_data.get('assigned_to')
            )
        
        # Exportar si se solicitó
        if export_format and report_data:
            if export_format == 'excel':
                return export_to_excel(report_data, report_title, report_type)
            elif export_format == 'pdf':
                return export_to_pdf(report_data, report_title, report_type)
    
    context = {
        'form': form,
        'report_data': report_data,
        'report_type': report_type,
        'report_title': report_title,
    }
    return render(request, 'reports/report.html', context)


# ==================== REPORTES ====================

def get_sales_report(start_date=None, end_date=None):
    """Reporte de ventas: ingresos, tendencias, etc."""
    title = "📊 Reporte de Ventas"
    
    opportunities = Opportunity.objects.filter(status='ganada')
    
    if start_date:
        opportunities = opportunities.filter(created_at__gte=start_date)
    if end_date:
        opportunities = opportunities.filter(created_at__lte=end_date)
    
    total_value = opportunities.aggregate(Sum('amount'))['amount__sum'] or 0
    total_deals = opportunities.count()
    
    # Por vendedor
    by_seller = opportunities.values('assigned_to__username').annotate(
        count=Count('id'),
        total=Sum('amount'),
        avg=Sum('amount') / Count('id')
    ).order_by('-total')
    
    # Por mes
    from django.db.models.functions import TruncMonth
    by_month = opportunities.annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-month')
    
    report_data = {
        'title': title,
        'summary': {
            'total_value': float(total_value),
            'total_deals': total_deals,
            'average_deal': float(total_value / total_deals) if total_deals > 0 else 0,
        },
        'by_seller': list(by_seller),
        'by_month': list(by_month),
    }
    
    return report_data, title


def get_customers_report(start_date=None, end_date=None, seller=None):
    """Reporte de clientes: estado, actividad, valor, etc."""
    title = "👥 Reporte de Clientes"
    
    customers = Customer.objects.all()
    
    if seller:
        customers = customers.filter(assigned_to=seller)
    
    total_customers = customers.count()
    active_customers = customers.filter(is_active=True).count()
    inactive_customers = customers.filter(is_active=False).count()
    
    # Clientes con más oportunidades
    top_customers = customers.annotate(
        opp_count=Count('opportunities'),
        opp_value=Sum('opportunities__amount')
    ).filter(opp_count__gt=0).order_by('-opp_value')[:10]
    
    # Clientes sin actividad
    dormant_customers = customers.annotate(
        opp_count=Count('opportunities'),
        followup_count=Count('followups')
    ).filter(opp_count=0, followup_count=0)
    
    report_data = {
        'title': title,
        'summary': {
            'total': total_customers,
            'active': active_customers,
            'inactive': inactive_customers,
            'dormant': dormant_customers.count(),
        },
        'top_customers': list(top_customers.values('name', 'assigned_to__username', 'opp_count', 'opp_value')),
        'dormant_customers': list(dormant_customers.values('name', 'email', 'phone')),
    }
    
    return report_data, title


def get_sellers_report(start_date=None, end_date=None):
    """Reporte de vendedores: rendimiento, métricas, etc."""
    title = "🎯 Análisis de Vendedores"
    
    team_members = TeamMember.objects.filter(is_active_seller=True)
    
    sellers_data = []
    for member in team_members:
        opportunities = Opportunity.objects.filter(assigned_to=member.user)
        
        if start_date:
            opportunities = opportunities.filter(created_at__gte=start_date)
        if end_date:
            opportunities = opportunities.filter(created_at__lte=end_date)
        
        won = opportunities.filter(status='ganada')
        
        seller_info = {
            'name': member.user.get_full_name() or member.user.username,
            'email': member.user.email,
            'department': member.department,
            'total_opportunities': opportunities.count(),
            'won_opportunities': won.count(),
            'lost_opportunities': opportunities.filter(status='perdida').count(),
            'open_opportunities': opportunities.filter(status__in=['abierta', 'calificada', 'propuesta', 'negociacion']).count(),
            'pipeline_value': float(opportunities.filter(status__in=['abierta', 'calificada', 'propuesta', 'negociacion']).aggregate(Sum('amount'))['amount__sum'] or 0),
            'won_value': float(won.aggregate(Sum('amount'))['amount__sum'] or 0),
            'win_rate': (won.count() / opportunities.count() * 100) if opportunities.count() > 0 else 0,
            'avg_deal_size': float(opportunities.aggregate(Sum('amount'))['amount__sum'] or 0 / opportunities.count()) if opportunities.count() > 0 else 0,
            'customers': member.user.customers.count(),
            'followups': Followup.objects.filter(user=member.user).count(),
        }
        sellers_data.append(seller_info)
    
    sellers_data.sort(key=lambda x: x['won_value'], reverse=True)
    
    report_data = {
        'title': title,
        'sellers': sellers_data,
    }
    
    return report_data, title


def get_followups_report(start_date=None, end_date=None, seller=None):
    """Reporte de seguimientos: actividad, tipos, estado, etc."""
    title = "💬 Reporte de Seguimientos"
    
    followups = Followup.objects.all()
    
    if seller:
        followups = followups.filter(user=seller)
    
    if start_date:
        followups = followups.filter(date__gte=start_date)
    if end_date:
        followups = followups.filter(date__lte=end_date)
    
    total_followups = followups.count()
    pending = followups.filter(status='pendiente').count()
    completed = followups.filter(status='completado').count()
    overdue = followups.filter(status='vencido').count()
    
    # Por tipo
    by_type = followups.values('type').annotate(count=Count('id')).order_by('-count')
    
    # Por vendedor
    by_seller = followups.values('user__username').annotate(
        count=Count('id'),
        completed_count=Count('id', filter=Q(status='completado'))
    ).order_by('-count')
    
    report_data = {
        'title': title,
        'summary': {
            'total': total_followups,
            'pending': pending,
            'completed': completed,
            'overdue': overdue,
        },
        'by_type': list(by_type),
        'by_seller': list(by_seller),
    }
    
    return report_data, title


def get_opportunities_report(start_date=None, end_date=None, customer=None, seller=None, statuses=None):
    """Reporte de oportunidades: detallado con filtros"""
    title = "📈 Reporte de Oportunidades"
    
    opportunities = Opportunity.objects.all()
    
    if start_date:
        opportunities = opportunities.filter(created_at__gte=start_date)
    if end_date:
        opportunities = opportunities.filter(created_at__lte=end_date)
    if customer:
        opportunities = opportunities.filter(customer=customer)
    if seller:
        opportunities = opportunities.filter(assigned_to=seller)
    if statuses:
        opportunities = opportunities.filter(status__in=statuses)
    
    # Resumen
    total_opportunities = opportunities.count()
    total_value = opportunities.aggregate(Sum('amount'))['amount__sum'] or 0
    
    by_status = opportunities.values('status').annotate(
        count=Count('id'),
        value=Sum('amount')
    ).order_by('-value')
    
    by_priority = opportunities.values('priority').annotate(
        count=Count('id'),
        value=Sum('amount')
    )
    
    # Oportunidades por vencer
    today = timezone.now().date()
    due_soon = opportunities.filter(
        expected_close_date__gte=today,
        expected_close_date__lte=today + timezone.timedelta(days=7),
        status__in=['abierta', 'calificada', 'propuesta', 'negociacion']
    )
    
    opportunities_list = opportunities.values(
        'title', 'customer__name', 'assigned_to__username',
        'amount', 'status', 'priority', 'expected_close_date'
    ).order_by('-amount')
    
    report_data = {
        'title': title,
        'summary': {
            'total': total_opportunities,
            'total_value': float(total_value),
            'avg_value': float(total_value / total_opportunities) if total_opportunities > 0 else 0,
            'due_soon': due_soon.count(),
        },
        'by_status': list(by_status),
        'by_priority': list(by_priority),
        'opportunities': list(opportunities_list),
    }
    
    return report_data, title


def get_pipeline_report(start_date=None, end_date=None, seller=None):
    """Reporte de pipeline: análisis de flujo"""
    title = "🔄 Análisis de Pipeline"
    
    opportunities = Opportunity.objects.filter(
        status__in=['abierta', 'calificada', 'propuesta', 'negociacion']
    )
    
    if seller:
        opportunities = opportunities.filter(assigned_to=seller)
    
    if start_date:
        opportunities = opportunities.filter(created_at__gte=start_date)
    if end_date:
        opportunities = opportunities.filter(created_at__lte=end_date)
    
    # Pipeline por etapa
    abierta_value = opportunities.filter(status='abierta').aggregate(Sum('amount'))['amount__sum'] or 0
    calificada_value = opportunities.filter(status='calificada').aggregate(Sum('amount'))['amount__sum'] or 0
    propuesta_value = opportunities.filter(status='propuesta').aggregate(Sum('amount'))['amount__sum'] or 0
    negociacion_value = opportunities.filter(status='negociacion').aggregate(Sum('amount'))['amount__sum'] or 0
    
    pipeline_stages = {
        'abierta': float(abierta_value),
        'calificada': float(calificada_value),
        'propuesta': float(propuesta_value),
        'negociacion': float(negociacion_value),
    }
    
    total_pipeline = sum(pipeline_stages.values())
    
    # Por vendedor
    by_seller = opportunities.values('assigned_to__username').annotate(
        value=Sum('amount'),
        count=Count('id')
    ).order_by('-value')
    
    # Convertir a float
    by_seller_list = []
    for seller_data in by_seller:
        seller_data['value'] = float(seller_data['value'] or 0)
        by_seller_list.append(seller_data)
    
    # Probabilidad ponderada
    probabilities = {
        'abierta': 0.1,
        'calificada': 0.3,
        'propuesta': 0.6,
        'negociacion': 0.8,
    }
    
    weighted_value = (
        float(abierta_value) * probabilities['abierta'] +
        float(calificada_value) * probabilities['calificada'] +
        float(propuesta_value) * probabilities['propuesta'] +
        float(negociacion_value) * probabilities['negociacion']
    )
    
    report_data = {
        'title': title,
        'pipeline_stages': pipeline_stages,
        'total_pipeline': total_pipeline,
        'weighted_forecast': weighted_value,
        'by_seller': by_seller_list,
    }
    
    return report_data, title



def export_to_excel(report_data, title, report_type):
    """Exportar reporte a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse("Error: openpyxl no está instalado. Ejecuta: pip install openpyxl", status=500)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Encabezado
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    
    row = 4
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if report_type == 'sales':
        ws[f'A{row}'] = "RESUMEN DE VENTAS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Valor Total:"
        ws[f'B{row}'] = report_data['summary']['total_value']
        row += 1
        ws[f'A{row}'] = "Total de Deals:"
        ws[f'B{row}'] = report_data['summary']['total_deals']
        row += 1
        ws[f'A{row}'] = "Deal Promedio:"
        ws[f'B{row}'] = report_data['summary']['average_deal']
        row += 2
        
        ws[f'A{row}'] = "Vendedor"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Total"
        ws[f'D{row}'] = "Promedio"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        row += 1
        
        for seller in report_data['by_seller']:
            ws[f'A{row}'] = seller['assigned_to__username']
            ws[f'B{row}'] = seller['count']
            ws[f'C{row}'] = float(seller['total'])
            ws[f'D{row}'] = float(seller['avg'])
            row += 1
    
    elif report_type == 'customers':
        ws[f'A{row}'] = "RESUMEN DE CLIENTES"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        for key, value in report_data['summary'].items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        ws[f'A{row}'] = "CLIENTES PRINCIPALES"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Cliente"
        ws[f'B{row}'] = "Vendedor"
        ws[f'C{row}'] = "Oportunidades"
        ws[f'D{row}'] = "Valor Total"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        row += 1
        
        for customer in report_data.get('top_customers', []):
            ws[f'A{row}'] = customer['name']
            ws[f'B{row}'] = customer.get('assigned_to__username', 'N/A')
            ws[f'C{row}'] = customer['opp_count']
            ws[f'D{row}'] = float(customer['opp_value'] or 0)
            row += 1
        
        row += 2
        ws[f'A{row}'] = "CLIENTES SIN ACTIVIDAD"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Cliente"
        ws[f'B{row}'] = "Email"
        ws[f'C{row}'] = "Teléfono"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
        
        row += 1
        
        for customer in report_data.get('dormant_customers', []):
            ws[f'A{row}'] = customer['name']
            ws[f'B{row}'] = customer.get('email', 'N/A')
            ws[f'C{row}'] = customer.get('phone', 'N/A')
            row += 1
    
    elif report_type == 'sellers':
        ws[f'A{row}'] = "ANÁLISIS DE VENDEDORES"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 2
        
        headers = ['Nombre', 'Email', 'Oportunidades', 'Ganadas', 'Perdidas', 'Pipeline', 'Valor Ganado', 'Win Rate %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        row += 1
        
        for seller in report_data.get('sellers', []):
            ws[f'A{row}'] = seller['name']
            ws[f'B{row}'] = seller['email']
            ws[f'C{row}'] = seller['total_opportunities']
            ws[f'D{row}'] = seller['won_opportunities']
            ws[f'E{row}'] = seller['lost_opportunities']
            ws[f'F{row}'] = float(seller['pipeline_value'])
            ws[f'G{row}'] = float(seller['won_value'])
            ws[f'H{row}'] = round(seller['win_rate'], 2)
            row += 1
    
    elif report_type == 'followups':
        ws[f'A{row}'] = "RESUMEN DE SEGUIMIENTOS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total:"
        ws[f'B{row}'] = report_data['summary']['total']
        row += 1
        ws[f'A{row}'] = "Pendientes:"
        ws[f'B{row}'] = report_data['summary']['pending']
        row += 1
        ws[f'A{row}'] = "Completados:"
        ws[f'B{row}'] = report_data['summary']['completed']
        row += 1
        ws[f'A{row}'] = "Vencidos:"
        ws[f'B{row}'] = report_data['summary']['overdue']
        row += 2
        
        ws[f'A{row}'] = "POR TIPO"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Tipo"
        ws[f'B{row}'] = "Cantidad"
        
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        row += 1
        
        for item in report_data.get('by_type', []):
            type_map = {'llamada': 'Llamada', 'email': 'Email', 'reunion': 'Reunión', 'nota': 'Nota'}
            ws[f'A{row}'] = type_map.get(item['type'], item['type'])
            ws[f'B{row}'] = item['count']
            row += 1
        
        row += 2
        ws[f'A{row}'] = "POR VENDEDOR"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Vendedor"
        ws[f'B{row}'] = "Total"
        ws[f'C{row}'] = "Completados"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        row += 1
        
        for seller in report_data.get('by_seller', []):
            ws[f'A{row}'] = seller['user__username']
            ws[f'B{row}'] = seller['count']
            ws[f'C{row}'] = seller['completed_count']
            row += 1
    
    elif report_type == 'opportunities':
        ws[f'A{row}'] = "RESUMEN DE OPORTUNIDADES"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total:"
        ws[f'B{row}'] = report_data['summary']['total']
        row += 1
        ws[f'A{row}'] = "Valor Total:"
        ws[f'B{row}'] = float(report_data['summary']['total_value'])
        row += 1
        ws[f'A{row}'] = "Valor Promedio:"
        ws[f'B{row}'] = float(report_data['summary']['avg_value'])
        row += 2
        
        ws[f'A{row}'] = "Título"
        ws[f'B{row}'] = "Cliente"
        ws[f'C{row}'] = "Vendedor"
        ws[f'D{row}'] = "Monto"
        ws[f'E{row}'] = "Estado"
        ws[f'F{row}'] = "Prioridad"
        ws[f'G{row}'] = "Vencimiento"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        row += 1
        
        for opp in report_data.get('opportunities', []):
            ws[f'A{row}'] = opp['title']
            ws[f'B{row}'] = opp['customer__name']
            ws[f'C{row}'] = opp['assigned_to__username']
            ws[f'D{row}'] = float(opp['amount'])
            ws[f'E{row}'] = opp['status']
            ws[f'F{row}'] = opp['priority']
            ws[f'G{row}'] = str(opp['expected_close_date'])
            row += 1
    
    elif report_type == 'pipeline':
        ws[f'A{row}'] = "ANÁLISIS DE PIPELINE"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Pipeline Total:"
        ws[f'B{row}'] = float(report_data['total_pipeline'])
        row += 1
        ws[f'A{row}'] = "Pronóstico Ponderado:"
        ws[f'B{row}'] = float(report_data['weighted_forecast'])
        row += 2
        
        ws[f'A{row}'] = "POR ETAPA"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Etapa"
        ws[f'B{row}'] = "Valor"
        ws[f'C{row}'] = "Porcentaje"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        row += 1
        
        for stage, value in report_data.get('pipeline_stages', {}).items():
            percentage = (value / report_data['total_pipeline'] * 100) if report_data['total_pipeline'] > 0 else 0
            ws[f'A{row}'] = stage.title()
            ws[f'B{row}'] = float(value)
            ws[f'C{row}'] = round(percentage, 2)
            row += 1
        
        row += 2
        ws[f'A{row}'] = "POR VENDEDOR"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Vendedor"
        ws[f'B{row}'] = "Valor"
        ws[f'C{row}'] = "Cantidad"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        row += 1
        
        for seller in report_data.get('by_seller', []):
            ws[f'A{row}'] = seller['assigned_to__username']
            ws[f'B{row}'] = float(seller['value'])
            ws[f'C{row}'] = seller['count']
            row += 1
    
    # Ajustar ancho de columnas (CORREGIDO)
    try:
        for column_cells in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column_cells:
                try:
                    # Solo procesar celdas normales, no fusionadas
                    if cell.data_type != 'f':
                        column_letter = cell.column_letter
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    except Exception as e:
        logger.warning(f"Error ajustando ancho de columnas: {e}")
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_{report_type}_{timezone.now().strftime("%d%m%Y")}.xlsx"'
    wb.save(response)
    
    logger.info(f"Reporte {report_type} exportado a Excel")
    
    return response

def export_to_pdf(report_data, title, report_type):
    """Exportar reporte a PDF - VERSIÓN COMPLETA Y CORREGIDA"""
    
    # Importar DENTRO de la función para evitar errores de módulo
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        logger.info(f"[OK] ReportLab imports successful")
    except ImportError as e:
        logger.error(f"[ERROR] ReportLab ImportError: {str(e)}", exc_info=True)
        return HttpResponse(f"Error: reportlab no esta instalado. {str(e)}", status=500)
    except Exception as e:
        logger.error(f"[ERROR] Unexpected error in imports: {str(e)}", exc_info=True)
        return HttpResponse(f"Error inesperado: {str(e)}", status=500)
    
    try:
        # Validar datos
        if not report_data:
            logger.warning("report_data is empty")
            return HttpResponse("Error: No hay datos para exportar", status=400)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{report_type}_{timezone.now().strftime("%d%m%Y")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=20,
            alignment=1,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph(title, title_style))
        
        # Fecha
        fecha_style = ParagraphStyle(
            'CustomDate',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=0
        )
        elements.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", fecha_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Contenido según tipo
        if report_type == 'sales':
            data = [['Métrica', 'Valor']]
            summary = report_data.get('summary', {})
            data.append(['Valor Total', f"${float(summary.get('total_value', 0) or 0):,.2f}"])
            data.append(['Total de Deals', str(summary.get('total_deals', 0))])
            data.append(['Deal Promedio', f"${float(summary.get('average_deal', 0) or 0):,.2f}"])
            
            table = Table(data, colWidths=[3 * inch, 3 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
            # Ventas por vendedor
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Ventas por Vendedor", styles['Heading3']))
            elements.append(Spacer(1, 0.1 * inch))
            
            data = [['Vendedor', 'Cantidad', 'Total', 'Promedio']]
            for seller in report_data.get('by_seller', []):
                data.append([
                    seller.get('assigned_to__username', 'N/A'),
                    str(seller.get('count', 0)),
                    f"${float(seller.get('total', 0) or 0):,.2f}",
                    f"${float(seller.get('avg', 0) or 0):,.2f}"
                ])
            
            if len(data) > 1:
                table = Table(data, colWidths=[2 * inch, 1.2 * inch, 1.5 * inch, 1.5 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
        
        elif report_type == 'customers':
            # Resumen
            data = [['Métrica', 'Cantidad']]
            for key, value in report_data.get('summary', {}).items():
                data.append([key.replace('_', ' ').title(), str(value)])
            
            table = Table(data, colWidths=[3 * inch, 2 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
            # Clientes principales
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Clientes Principales", styles['Heading3']))
            elements.append(Spacer(1, 0.1 * inch))
            
            if report_data.get('top_customers'):
                data = [['Cliente', 'Vendedor', 'Opor.', 'Valor']]
                for customer in report_data['top_customers']:
                    data.append([
                        customer.get('name', 'N/A')[:25],
                        customer.get('assigned_to__username', 'N/A'),
                        str(customer.get('opp_count', 0)),
                        f"${float(customer.get('opp_value', 0) or 0):,.0f}"
                    ])
                
                table = Table(data, colWidths=[1.5 * inch, 1.5 * inch, 1 * inch, 1.5 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
            
            # Clientes sin actividad
            if report_data.get('dormant_customers'):
                elements.append(Spacer(1, 0.3 * inch))
                elements.append(Paragraph("Clientes Sin Actividad", styles['Heading3']))
                elements.append(Spacer(1, 0.1 * inch))
                
                data = [['Cliente', 'Email', 'Teléfono']]
                for customer in report_data['dormant_customers'][:10]:
                    data.append([
                        customer.get('name', 'N/A')[:25],
                        customer.get('email', 'N/A'),
                        customer.get('phone', 'N/A')
                    ])
                
                table = Table(data, colWidths=[2 * inch, 2 * inch, 1.5 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
        
        elif report_type == 'sellers':
            data = [['Vendedor', 'Opor.', 'Ganadas', 'Pipeline', 'Win Rate', 'Promedio']]
            for seller in report_data.get('sellers', [])[:15]:
                data.append([
                    seller.get('name', 'N/A')[:20],
                    str(seller.get('total_opportunities', 0)),
                    str(seller.get('won_opportunities', 0)),
                    f"${float(seller.get('pipeline_value', 0) or 0):,.0f}",
                    f"{float(seller.get('win_rate', 0) or 0):.1f}%",
                    f"${float(seller.get('avg_deal_size', 0) or 0):,.0f}"
                ])
            
            table = Table(data, colWidths=[1.2 * inch, 0.7 * inch, 0.7 * inch, 1.2 * inch, 0.7 * inch, 0.8 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
        
        elif report_type == 'followups':
            # Resumen
            data = [['Estado', 'Cantidad']]
            summary = report_data.get('summary', {})
            data.append(['Total', str(summary.get('total', 0))])
            data.append(['Pendientes', str(summary.get('pending', 0))])
            data.append(['Completados', str(summary.get('completed', 0))])
            data.append(['Vencidos', str(summary.get('overdue', 0))])
            
            table = Table(data, colWidths=[3 * inch, 2 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
            # Por tipo
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Seguimientos por Tipo", styles['Heading3']))
            elements.append(Spacer(1, 0.1 * inch))
            
            type_map = {'llamada': 'Llamada', 'email': 'Email', 'reunion': 'Reunión', 'nota': 'Nota'}
            data = [['Tipo', 'Cantidad']]
            for item in report_data.get('by_type', []):
                data.append([type_map.get(item.get('type', 'otro'), item.get('type', 'Otro')), str(item.get('count', 0))])
            
            if len(data) > 1:
                table = Table(data, colWidths=[2.5 * inch, 2.5 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
            
            # Por vendedor
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Actividad por Vendedor", styles['Heading3']))
            elements.append(Spacer(1, 0.1 * inch))
            
            data = [['Vendedor', 'Total', 'Completados', 'Tasa %']]
            for seller in report_data.get('by_seller', []):
                total = seller.get('count', 0)
                completed = seller.get('completed_count', 0)
                tasa = (completed / total * 100) if total > 0 else 0
                data.append([
                    seller.get('user__username', 'N/A')[:20],
                    str(total),
                    str(completed),
                    f"{tasa:.1f}%"
                ])
            
            if len(data) > 1:
                table = Table(data, colWidths=[1.8 * inch, 1 * inch, 1.2 * inch, 1 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
        
        elif report_type == 'opportunities':
            # Resumen
            data = [['Métrica', 'Valor']]
            summary = report_data.get('summary', {})
            data.append(['Total', str(summary.get('total', 0))])
            data.append(['Valor Total', f"${float(summary.get('total_value', 0) or 0):,.0f}"])
            data.append(['Promedio', f"${float(summary.get('avg_value', 0) or 0):,.0f}"])
            data.append(['Vencen Pronto', str(summary.get('due_soon', 0))])
            
            table = Table(data, colWidths=[3 * inch, 2 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
            # Detalle
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Detalle de Oportunidades", styles['Heading3']))
            elements.append(Spacer(1, 0.1 * inch))
            
            data = [['Título', 'Cliente', 'Monto', 'Estado', 'Vencimiento']]
            for opp in report_data.get('opportunities', [])[:15]:
                data.append([
                    opp.get('title', 'N/A')[:18],
                    opp.get('customer__name', 'N/A')[:15],
                    f"${float(opp.get('amount', 0) or 0):,.0f}",
                    opp.get('status', 'N/A')[:8],
                    str(opp.get('expected_close_date', ''))[:10]
                ])
            
            if len(data) > 1:
                table = Table(data, colWidths=[1.3 * inch, 1.3 * inch, 1 * inch, 1 * inch, 1 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
        
        elif report_type == 'pipeline':
            # Resumen
            data = [['Métrica', 'Valor']]
            total_pipeline = float(report_data.get('total_pipeline', 0) or 0)
            data.append(['Pipeline Total', f"${total_pipeline:,.0f}"])
            data.append(['Pronóstico Ponderado', f"${float(report_data.get('weighted_forecast', 0) or 0):,.0f}"])
            
            table = Table(data, colWidths=[3 * inch, 2 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
            # Por etapa
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Pipeline por Etapa", styles['Heading3']))
            elements.append(Spacer(1, 0.1 * inch))
            
            data = [['Etapa', 'Valor', '%']]
            for stage, value in report_data.get('pipeline_stages', {}).items():
                percentage = (value / total_pipeline * 100) if total_pipeline > 0 else 0
                data.append([
                    stage.title(),
                    f"${float(value or 0):,.0f}",
                    f"{percentage:.1f}%"
                ])
            
            if len(data) > 1:
                table = Table(data, colWidths=[1.5 * inch, 2 * inch, 1 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
            
            # Por vendedor
            elements.append(Spacer(1, 0.3 * inch))
            elements.append(Paragraph("Pipeline por Vendedor", styles['Heading3']))
            elements.append(Spacer(1, 0.1 * inch))
            
            if report_data.get('by_seller'):
                data = [['Vendedor', 'Valor', 'Oportunidades', '%']]
                for seller in report_data['by_seller']:
                    seller_value = float(seller.get('value', 0) or 0)
                    seller_percentage = (seller_value / total_pipeline * 100) if total_pipeline > 0 else 0
                    data.append([
                        seller.get('assigned_to__username', 'N/A')[:20],
                        f"${seller_value:,.0f}",
                        str(seller.get('count', 0)),
                        f"{seller_percentage:.1f}%"
                    ])
                
                table = Table(data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 0.8 * inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        logger.info(f"✓ Reporte {report_type} exportado a PDF correctamente")
        
        return response
    
    except KeyError as e:
        logger.error(f"✗ KeyError en export_to_pdf [{report_type}]: {str(e)}", exc_info=True)
        return HttpResponse(f"Error: Clave faltante en datos - {str(e)}", status=500)
    
    except Exception as e:
        logger.error(f"✗ Error inesperado en export_to_pdf [{report_type}]: {str(e)}", exc_info=True)
        return HttpResponse(f"Error al generar PDF: {str(e)}", status=500)
